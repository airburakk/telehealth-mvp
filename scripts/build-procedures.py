# -*- coding: utf-8 -*-
"""
İŞLEM KATALOĞU ÜRETİCİ — src/data/procedures.json
=================================================
Kaynak: T.C. Sağlık Bakanlığı "Kamu Sağlık Hizmetleri Fiyat Tarifesi" (EK-2 / KSHFT) .xlsx
(tek sayfa "KSHFT"; sütunlar: SIRA NO | KODU | İŞLEM ADI | AÇIKLAMALAR | FİYAT)

Tarife hiyerarşik: başlık satırları (KODU boş, bold) + kodlu işlem satırları.
Excel'de seviye bilgisi YOK (tüm başlıklar bold/indent 0). Bu yüzden her işlem,
KENDİSİNDEN ÖNCEKİ başlıkları EN YAKINDAN UZAĞA tarayıp İLK eşleşen branşa atanır
(nearest-header-first). Tanı/lab/radyoloji/patoloji/idari kalemler -> "others".

Kullanım:  python scripts/build-procedures.py <tarife.xlsx>
Gerekli:   pip install openpyxl
Çıktı:     src/data/procedures.json   (compact tuple format)

Tarife güncellendiğinde bu script yeniden çalıştırılır. HEADER_MAP klinik eşlemeyi
içerir; yeni başlık kalıpları çıkarsa buraya eklenir.
"""
import sys, os, json, datetime
import openpyxl

def norm(s):
    if not s: return ""
    s = s.replace("İ","i").replace("I","ı").lower()
    for a,b in [("ç","c"),("ğ","g"),("ı","i"),("ö","o"),("ş","s"),("ü","u"),("â","a"),("î","i"),("û","u")]:
        s = s.replace(a,b)
    return s

BRANCHES = [
    ("onkoloji","Onkoloji"),("kardiyoloji","Kardiyoloji"),("ortopedi","Ortopedi"),
    ("norosirurji","Nöroşirürji"),("sac-ekimi","Saç Ekimi"),("estetik","Estetik Cerrahi"),
    ("ivf","Tüp Bebek (IVF)"),("dis","Diş Tedavisi"),("goz","Göz Cerrahisi"),
    ("genel-cerrahi","Genel Cerrahi"),("dahiliye","Dahiliye (İç Hastalıkları)"),
    ("noroloji","Nöroloji"),("gastroenteroloji","Gastroenteroloji"),
    ("endokrinoloji","Endokrinoloji ve Metabolizma"),("nefroloji","Nefroloji"),
    ("gogus-hastaliklari","Göğüs Hastalıkları"),("hematoloji","Hematoloji"),
    ("romatoloji","Romatoloji"),("enfeksiyon","Enfeksiyon Hastalıkları"),
    ("dermatoloji","Dermatoloji (Cilt Hastalıkları)"),("psikiyatri","Psikiyatri"),
    ("fizik-tedavi","Fiziksel Tıp ve Rehabilitasyon"),("cocuk-sagligi","Çocuk Sağlığı ve Hastalıkları"),
    ("uroloji","Üroloji"),("kbb","Kulak Burun Boğaz (KBB)"),("kadin-dogum","Kadın Hastalıkları ve Doğum"),
    ("kvc","Kalp ve Damar Cerrahisi"),("gogus-cerrahisi","Göğüs Cerrahisi"),
    ("organ-nakli","Organ Nakli"),("radyasyon-onkolojisi","Radyasyon Onkolojisi"),
]
LABELS = dict(BRANCHES)
LABELS["others"] = "Diğer (Sınıflandırılmamış)"

# (branş, [tek bir başlıkta aranan normalize alt-dizeler]) — sıra önemli (spesifik/cerrahi önce)
HEADER_MAP = [
    ("sac-ekimi", ["sacli deri"]),
    ("dis", ["dis tedavileri","tedavi ve endodonti","endodonti","pedodonti","ortodonti","periodontoloji",
             "agiz, dis ve cene","agiz dis ve cene","dis paket","dis hekimi","teshis ve tedavi planlamasi","protez"]),
    ("radyasyon-onkolojisi", ["radyasyon onkolojisi","radyoterapi","brakiterapi","eksternal radyoterapi",
             "lineer akselerator","hipertermi","klinik onkolojik degerlendirme","radyoterapi tasarimi",
             "radyoterapi planlama","medikal radyasyon fizigi","portal goruntuleme","after-loading"]),
    ("organ-nakli", ["organ transplantasyonu","organ nakli"]),
    ("kardiyoloji", ["klinik kardiyoloji","elektrokardiyografi","ekokardiyografi","tanisal kalp kateterizasyonu",
             "tedavi amacli kalp kateterizasyonu","elektrofizyolojik calisma","kalp pili (pacemaker) ve icd",
             "dogumsal kalp hastaliklari"]),
    ("kvc", ["kardiyovaskuler sistem","kalp tumorleri","kalp pili","kardioverter","kalp ve buyuk damar",
             "kalp kapaklari","koroner arter","septal defekt","sinus valsalva","aortik anomali",
             "torasik aort anevrizmasi","pulmoner arter","direkt anevrizma","by-pass greft","insitu ven",
             "trombektomi","trombendarterektomi","trombendarter","venoz rekonstruksiyon","arteriovenoz fistul",
             "fistul disinda damar","ligasyon ve diger islemler","transpozisyonu","trunkus arteriosus",
             "total anormal pulmoner","perikard","shunt islemleri","intravaskuler kanulasyon",
             "transkateter tedavisi","arteryel embolektomi","venoz trombektomi"]),
    ("gogus-cerrahisi", ["solunum sistemi cerrahisi","trakea ve bron","akcigerler ve plevra","toraks duvari","diyafragma cerrahisi"]),
    ("genel-cerrahi", ["sindirim sistemi cerrahisi","karaciger","safra yollari","pankreas","ozefagus","mide-duodenum",
             "jejunum","appendiks","kolon","rektum","anus","govde ve karin on duvari","herniler","periton boslugu",
             "retroperiton","endokrin sistem cerrahisi","dalak","robotik cerrahi","govdede yapilan cerrahiler"]),
    ("kbb", ["bas ve boyun cerrahisi","boyun ve larinks","paranazal sinusler","agiz, dudak, dil",
             "salgi bezlerine yonelik","kulak ve kulak bolgesinin cerrahisi","ses ve isitme","burun","endoskopi"]),
    ("norosirurji", ["sinir sistemi cerrahisi","kraniyal cerrahi","kafa travma","kitle ve vaskuler ameliyat",
             "epilepsi ameliyat","konjenital spinal","intradural","disk cerrahisi","eksizyon ve dekompresyon",
             "fasiyal parali","periferik sinir cerrahisi","stereotaktik ve fonksiyonel","omurga cerrahisi",
             "spinal enstrumantasyon","vertebra enfeksiyon","orta kafa cukuru","retro labirenter"]),
    ("goz", ["goz ve adneksleri","periokuler","gozyasi drenaj","sasilik ve pediyatrik oftalmoloji",
             "konjonktiva-kornea","refraktif cerrahi","iris ve lens","glokom","retina-vitreus","orbita-okuler"]),
    ("uroloji", ["uriner sistem cerrahisi","bobrek","ureter","mesane","uretra","erkek genital sistemi",
             "penis","prostat","testis-epididim"]),
    ("ivf", ["invitro fertilizasyon","infertilite ameliyat","infertilite tetkik"]),
    ("kadin-dogum", ["kadin genital ve ureme","dogum islem","dogum ucretleri","gebelikte teshis","jinekoloji",
             "vajinal operasyon","abdominal operasyon","ozelligi olan operasyon","dogum oncesi tetkik"]),
    ("ortopedi", ["kemik ve eklem hastaliklari","alci","atel","traksiyon","kiriklarin","cikiklarin",
             "eklem mobilizasyonu","artroplasti","artrodez","osteomyelit","artroskopi","eklem acik cerrahi",
             "tekrarlayan cikik","pediyatrik ortopedi","pelvis ve kalca eklemi","eksternal fiksator","kemik uzatma",
             "ortopedik onkoloji","amputasyon","osteotomi","implant cikarma","kas iskelet sistemi",
             "el ve mikrocerrahi","konjenital anomaliler"]),
    ("gogus-hastaliklari", ["solunum sistemi"]),
    ("gastroenteroloji", ["sindirim sistemi"]),
    ("nefroloji", ["uriner sistem-nefroloji","nefroloji-diyaliz"]),
    ("hematoloji", ["hematoloji-onkoloji","aferez islemleri","kemik iligi nakil","kan bankasi","kan bilesenleri","kemik iligi nakli"]),
    ("psikiyatri", ["psikiyatrik calismalar","psikiyatri hizmetleri","uyku arastirma","uyku tetkik"]),
    ("noroloji", ["elektroensefalograf","elektromiyograf","uyarilmis potansiyel"]),
    ("fizik-tedavi", ["fizik tedavi","rehabilitasyon","hidroterapi","balneoterapi"]),
    ("cocuk-sagligi", ["yeni dogan uygulamalari","yenidogan destegi","yenidogan yogun bakim"]),
    ("dermatoloji", ["dermis ve epidermis","cilt bakimi","kimyasal soyma","peeling","roller uygulamasi","mezoterapi",
             "botulinum","dolgu maddesi","trombositten zengin plazma","epilasyon amacli","lazer ile deri",
             "deri genclestirme","stria ve skar","sellulit","dovme silinmesi","deri ve derialti","diger lazer",
             "lazer uygulamalari","minimal invaziv uygulamalar","deri"]),
    ("estetik", ["yuz ve boyun bolgesi","vucut hatlarinin","genital bolge","greftler","flepler","meme","goz cevresi","yuz ve boyun"]),
    ("others", ["saglik raporlari","seyahat sagligi","geleneksel, tamamlayici","geleneksel,tamamlayici","kayropraksi",
             "ozon uygulamasi","spor hekimligi","podoloji","su alti hekimligi","hiperbarik","yatak ucretleri",
             "hekim muayeneleri","genel uygulamalar-girisimler","kateter islemleri","ameliyathane","anestezi",
             "tani, tedavi ve yogun bakim","algoloji","agri tedavisi","sinir bloklari","sinir blogu","enjeksiyonlar",
             "noromodulasyon","noroplasti","norolitik","somatik sinir","sempatik sinir","intraspinal sinir",
             "yogun bakim","palyatif bakim","nukleer","radyonuklid","sintigraf","radyolojik goruntuleme","direkt grafi",
             "uzun kemik grafi","akciger grafi","bacak uzunluk","duz karin grafi","kafa grafi","telekardiogram",
             "pelvis grafi","vertebra grafi","kontrastli tetkik","anjiyografik tetkik","venografik","nororadyolojik",
             "kemik dansitometre","ultrasonografik","renkli doppler","bilgisayarli tomografi","manyetik rezonans",
             "artrografi","periferik anjiyografi","girisimsel radyolojik","laboratuvar islemleri","biyokimya laboratuvar",
             "alerji test","spesifik ige","aminoasitler","dinamik test","uyari testi","baskilama testi","tolerans testi",
             "mixed meal","jenerasyon testi","ilac duzeyi","monoklonal antikor","akim sitometresi","yasadisi ve kotuye",
             "tarama analizleri","dogrulama analizleri","molekuler mikrobiyoloji","sitogenetik","molekuler genetik",
             "onkolojik molekuler","patoloji","histopatolojik","sitolojik materyaller","otopsi","ozel patolojik",
             "elektron mikroskop","meslek hastaliklari","is sagligi ve guvenligi","referans laboratuvar",
             "mikrobiyolojik test","paraziter ve bakteriyel","immunolojik analiz","zoonotik","virolojik test",
             "tuberkuloz","toksikolojik analiz","adli tip","morg listesi","biyoloji listesi","toksikoloji listesi",
             "narkotik listesi","kannabinoid","afyon","kokain","her turlu ilac analizi","her cesit malzeme",
             "diger uyusturucu","gida ve cesitli","enstrumantal analiz","adli raporlar","ambulans","dolasim sistemi destegi",
             "hava yolu ve solunum destegi","dolasim destegi","stabilizasyon","check-up"]),
    # kaba (letter-section) yedekler
    ("dis", ["c- dis"]),
    ("genel-cerrahi", ["f- robotik"]),
    ("others", ["a- saglik","b- seyahat","g-ambulans","g- ambulans","h-adli","h- adli","i- check-up"]),
]
NAME_HAIR = ["sac ekimi","kas ekimi","kirpik ekimi","sakal ekimi"]
WINDOW = 12

def classify_header(h):
    hn = norm(h)
    for key, subs in HEADER_MAP:
        for s in subs:
            if s in hn:
                return key
    return None

def price_of(v):
    if v in (None, ""): return None
    try: return int(round(float(v)))
    except Exception: return None

def main():
    if len(sys.argv) < 2:
        print("Kullanım: python scripts/build-procedures.py <tarife.xlsx>"); sys.exit(1)
    path = sys.argv[1]
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True)):
        rows.append({
            "kod": (str(row[1]).strip() if row[1] not in (None,"") else ""),
            "ad": (str(row[2]).strip() if row[2] not in (None,"") else ""),
            "fiyat": row[4],
        })
    items = []
    stack = []
    for r in rows:
        if r["ad"] and not r["kod"]:
            stack.append(r["ad"])
            if len(stack) > WINDOW: stack = stack[-WINDOW:]
            continue
        if not r["kod"]:
            continue
        key = None
        for h in reversed(stack):
            key = classify_header(h)
            if key: break
        if not key and any(s in norm(r["ad"]) for s in NAME_HAIR): key = "sac-ekimi"
        if not key: key = "others"
        items.append([r["kod"], r["ad"], price_of(r["fiyat"]), key, stack[-1] if stack else ""])

    here = os.path.dirname(os.path.abspath(__file__))
    dest = os.path.join(here, "..", "src", "data", "procedures.json")
    os.makedirs(os.path.dirname(dest), exist_ok=True)
    out = {
        "generatedAt": datetime.date.today().isoformat(),
        "source": "T.C. Sağlık Bakanlığı — Kamu Sağlık Hizmetleri Fiyat Tarifesi (EK-2 KSHFT)",
        "note": "price = taban (resmi tarife, ₺). Tavan = taban × 3 (uygulamada hesaplanır).",
        "branchLabels": LABELS,
        "items": items,
    }
    with open(dest, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    from collections import Counter
    c = Counter(it[3] for it in items)
    print(f"{len(items)} işlem -> {os.path.relpath(dest)}")
    for k,_ in BRANCHES:
        print(f"  {c.get(k,0):5d}  {k}")
    print(f"  {c.get('others',0):5d}  others")

if __name__ == "__main__":
    main()
